from openpyxl import load_workbook


def sum_xl_nums(file: str) -> str:
    workbook = load_workbook(file, data_only=True)

    total_sum = 0

    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]

        for row in worksheet.iter_rows(values_only=True):
            for cell_value in row:
                if isinstance(cell_value, (int, float)):
                    total_sum += cell_value

    return f'Total Sum: {total_sum}'


if __name__ == '__main__':
    print(sum_xl_nums('./nums.xlsx'))
