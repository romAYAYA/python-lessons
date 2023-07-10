import sys

import openpyxl
from PyQt5.QtWidgets import QApplication, QMainWindow, QLabel, QLineEdit, QPushButton, QMessageBox
import sqlite3
from openpyxl import Workbook

conn = sqlite3.connect("data.db")
cur = conn.cursor()

cur.execute(
    "CREATE TABLE IF NOT EXISTS products (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT, quantity INTEGER, price REAL)")


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Data Import/Export")
        self.setGeometry(100, 100, 300, 200)

        self.path_label = QLabel("File Path:", self)
        self.path_label.setGeometry(20, 20, 80, 30)

        self.path_text_edit = QLineEdit(self)
        self.path_text_edit.setGeometry(100, 20, 150, 30)

        self.import_button = QPushButton("Import", self)
        self.import_button.setGeometry(50, 70, 80, 30)
        self.import_button.clicked.connect(self.import_data)

        self.export_button = QPushButton("Export", self)
        self.export_button.setGeometry(150, 70, 80, 30)
        self.export_button.clicked.connect(self.export_data)

    def import_data(self):
        file_path = self.path_text_edit.text()
        if not file_path:
            QMessageBox.warning(self, "Warning", "Please enter a file path.")
            return

        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=1, values_only=True):
                name, quantity, price = row
                cur.execute("INSERT INTO products (name, quantity, price) VALUES (?, ?, ?)", (name, quantity, price))
                conn.commit()

            QMessageBox.information(self, "Success", "Data imported successfully.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred: {str(e)}")

    def export_data(self):
        try:
            cur.execute("SELECT id, name, quantity, price FROM products")
            data = cur.fetchall()

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["ID", "Name", "Quantity", "Price"])

            for row in data:
                sheet.append(row)

            workbook.save("temp/exported_data.xlsx")
            QMessageBox.information(self, "Success", "Data exported successfully.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"An error occurred: {str(e)}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
