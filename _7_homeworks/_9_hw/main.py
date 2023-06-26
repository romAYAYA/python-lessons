import re
from collections import Counter
import openpyxl


# TODO task 1
# def extract_domain(email: str) -> str | None:
#     pattern = r'@([\w.-]+)'
#     match = re.search(pattern, email)
#     if match:
#         return match.group(1)
#     return None


# TODO task 2

# def extract_words_starting_with_vowel(text: str) -> list:
#     pattern = r'\b[аеёиоуыэюяaeiou]\w+'
#     matches = re.findall(pattern, text, flags=re.IGNORECASE | re.UNICODE)
#     return matches


# TODO task 3

# def divide_string(string: str, dividers: str) -> list[str]:
#     pattern = '|'.join(map(re.escape, dividers))
#     substrings = re.split(pattern, string)
#     return substrings


# TODO task 4

# def election():
#     candidates = ['Anime', 'Roma', 'Dima', 'Woman']
#
#     votes = []
#     while True:
#         candidate = input('Enter a candidate name (Anime, Roma, Dima, Woman) or "done" to finish: ')
#         if candidate == 'done':
#             break
#         if candidate in candidates:
#             votes.append(candidate)
#         else:
#             print("Invalid candidate name. Please try again.")
#
#     if not votes:
#         print("No votes recorded.")
#         return
#
#     vote_count = Counter(votes)
#     max_vote_count = max(vote_count.values())
#     winners = [candidate for candidate, count in vote_count.items() if count == max_vote_count]
#     winner = min(winners, key=lambda x: len(x))
#
#     print(f"\nElection Results: \nTotal votes cast: {len(votes)} \nCandidate\tVotes \n-----------------------")
#     for candidate in candidates:
#         print(candidate, "\t\t", vote_count[candidate])
#     print("-----------------------")
#     print(f"Winner: {winner}")


# TODO task 5

def read_xlsx_files(file1: str, file2: str, file3: str, output_file: str) -> str:
    data = []
    max_rows = 0

    for file in [file1, file2, file3]:
        wb = openpyxl.load_workbook(file)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        max_rows = max(max_rows, len(rows))
        data.append(rows)

    combined_wb = openpyxl.Workbook()
    combined_sheet = combined_wb.active

    for i, rows in enumerate(data):
        for j, row in enumerate(rows):
            for k, value in enumerate(row):
                combined_sheet.cell(row=j + 1, column=(i * 0) + k + 1, value=value)

    combined_wb.save(output_file)
    return f"Data written to {output_file} successfully."


if __name__ == '__main__':
    # TODO task 1
    # email = 'supernaturalmlady@gmail.com'
    # print(extract_domain(email))

    # TODO task 2
    # text = 'Урал, Ернар, море, Roma, Alla, Anime, аниме, aboba'
    # print(extract_words_starting_with_vowel(text))

    # TODO task 3
    # string = 'Hello-world! How;are:you?'
    # dividers = ['-', ';', ':']
    # print(divide_string(string, dividers))

    # TODO task 4
    # election()

    # TODO task 5
    print(read_xlsx_files('./temp/Лист1.xlsx', './temp/Лист2.xlsx', './temp/Лист3.xlsx', './temp/output.xlsx'))
