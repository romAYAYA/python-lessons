import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# Read
# Create
# Change (Read + Create)
# Template (уже есть красивый файл с внешним видом / с форматированием )

def read():
    workbook: Workbook = openpyxl.load_workbook('temp/data1.xlsx')
    worksheet: Worksheet = workbook.active  # todo активирует активный
    worksheet = workbook['Рабочий лист']  # todo активирует нужный рабочий лист

    cell1 = worksheet.cell(row=2, column=1).value
    cell2 = worksheet['B1'].value
    print(cell1)
    print(cell2)

    matrix = []
    for row_i in range(1, 4 + 1):
        local_list = []
        for column_i in range(1, 2 + 1):
            value = worksheet.cell(row=row_i, column=column_i).value
            local_list.append(value)
        matrix.append(local_list)
    print(matrix)

    print('\n\n\n\n')

    for j in matrix:
        print(j)


pass

if __name__ == '__main__':
    read()
