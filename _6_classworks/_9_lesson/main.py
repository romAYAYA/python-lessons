import openpyxl
import json


def read_and_write():
    workbook = openpyxl.load_workbook('temp/work.xlsx')
    worksheet = workbook['Sheet1']

    data_from_excel = []
    max_row = worksheet.max_row  # кривая штука (недоработка разработчиков)
    max_column = worksheet.max_column  # кривая штука (недоработка разработчиков)
    for row_i in range(1, max_row + 1):
        local_list = []
        for column_i in range(1, max_column + 1):
            value = worksheet.cell(row=row_i, column=column_i).value
            local_list.append(value)
        data_from_excel.append(local_list)

    with open('temp/equal.txt', mode='r', encoding='utf-8') as file:
        lines = file.readlines()
    clear_lines = []
    for line in lines:
        line = line.strip()  # Remove leading/trailing whitespace, including the newline character
        clear_lines.append(int(line))

    # сравнить
    res = []
    for j in data_from_excel:
        if j[0] in clear_lines:
            new_dict = {'value': j[0], 'degree ': j[1]}
            res.append(new_dict)

    with open('temp/output.json', mode='w', encoding='utf-8') as file:
        json.dump(res, file)


if __name__ == '__main__':
    read_and_write()
