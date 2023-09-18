import openpyxl


class Worker:
    PERSONAL_BONUS = 1.15

    def __repr__(self):
        return f'<Person {self.name} {self.salary} {self.position} {self.date_register}>'

    def __init__(self, name: str, position: str, salary: any, is_reserved: str, date_register):
        self.name: str = name
        self.position: str = position
        self.salary: float = float(salary)
        self.is_reserved: bool = str(is_reserved).strip().lower() == 'да'
        self.date_register = date_register

    def serialize_to_json(self) -> dict:
        return {'name': self.name, 'position': self.position, 'salary': self.salary,
                'is_reserved': self.is_reserved,
                'date_register': self.date_register}


workers = []
workbook = openpyxl.load_workbook('data.xlsx')
worksheet = workbook.active

for i in range(2, worksheet.max_row + 1):
    worker = Worker(
        name=worksheet.cell(row=i, column=1).value,
        position=worksheet.cell(row=i, column=2).value,
        salary=worksheet.cell(row=i, column=3).value,
        is_reserved=worksheet.cell(row=i, column=4).value,
        date_register=worksheet.cell(row=i, column=5).value
    )
    workers.append(worker)

only_reserved = list(filter(lambda x: x.is_reserved, workers))

print(only_reserved)











